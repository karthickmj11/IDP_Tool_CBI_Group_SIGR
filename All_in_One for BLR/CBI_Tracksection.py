import openpyxl
def check(s):
    if(s[:2]=="TC"):
        return s[2:]
    else:
        return s[1:]
def run(loc,op_l):
    f = open(op_l+'\\Tracksection.xml','w')
    wb = openpyxl.load_workbook(loc)
    sh = wb['TC']
    i=2
    u = 100
    sector = str(sh.cell(row=2,column=2).value)[-1:]
    f.write('''<?xml version="1.0" ?>
<ObjectPropertyModule Generation_Date="2019-02-25" Project="BLR" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="TrackSection.xsd">
  <Versions>
    <Version Name="ATS_SYSTEM_PARAMETERS#Comment" Value="XML file containing the System Data"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Date" Value="2011-03-19"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#SyPD_Version" Value="_none_"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#UEVOLtoIDP_Version" Value="2.1.7"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Writer_Name" Value="IconisDigesterCore 0.0.4091.29806"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#XML_File_Version" Value="1.6.6.1078"/>
    <Version Name="ICONIS_ATS_Equipment#SyID_Version" Value="Y3-64 A427945-J1"/>
    <Version Name="ICONIS_IDP#Customisation_SwRSAD_Version" Value="none"/>
    <Version Name="ICONIS_IDP#Product_SwRSAD_Version" Value="Y3-64 A427875-A"/>
    <Version Name="ICONIS_IDP#Product_Version" Value="10.3.4 (revision 3010)"/>
    <Version Name="ICONIS_IDP#Project_Version" Value="BLR#V10.3.4 (revision 3060)"/>
    <Version Name="Project#ATS_Custom_Reference_Database_Version" Value="0.0.3.565"/>
    <Version Name="Project#Database_Version" Value="0.0.3.565"/>
  </Versions>\n''')
    f.write('  <Classes>\n')
    f.write('    <Class name="TrackSection">\n')
    f.write('      <Objects>\n')
    while(sh.cell(row=i,column=2).value!=None):
        f.write('        <Object name="T'+check(str(sh.cell(row=i,column=2).value))+'" rules="update_or_create">\n')
        f.write('          <Properties>\n')
        f.write('            <Property dt="boolean" name="ATDasBoundary">false</Property>\n')
        f.write('            <Property dt="i4" name="CBTCArea">0</Property>\n')
        f.write('            <Property dt="string" name="CheckableStatus">0</Property>\n')
        f.write('            <Property dt="i4" name="ConsiderPointStraight">0</Property>\n')
        f.write('            <Property dt="i4" name="HILCWithConfirmation">1</Property>\n')
        f.write('            <Property dt="i4" name="HMITrainDirectionAutoResetTimer">0</Property>\n')
        f.write('            <Property dt="string" name="Interlocking">ASCV_'+sector+'</Property>\n')
        f.write('            <Property dt="string" name="ManagedKeys">ARSTPMA_TRAINLEAVEWITHOUTCONTROL;ATCM_ANSWER;ATCM_RELEASE_1;ATCM_RELEASE_1_TRACK;ATCM_RELEASE_2;ATCM_RELEASE_2_TRACK;ATCM_RESET;ATCM_RESET_TRACK;ATCM_STATE_ERROR;ATCM_STATUS_1;ATCM_STATUS_1_TRACK;ATCM_STATUS_2;ATCM_STATUS_2_TRACK;ATCM_WAIT_SAFETY_RELEASE;ATCM_WAIT_SAFETY_RELEASE_TRACK;ATCM_WAIT_SAFETY_STATUS;ATCM_WAIT_SAFETY_STATUS_TRACK;MSG_TDS_177_ATSPTIMISMATCH;MSG_TDS_177_PTIIDENTALREADYEXISTINATS;MSG_TDS_177_RECEIVEPTIMESSAGE;MSG_U200PTI_PERFORMSPTIINIT;MSG_U200PTI_UPDATE_DRIVERID;MSG_U200PTI_UPDATE_DRIVERID_MISMATCH;MSG_U200PTI_UPDATE_DRIVERID_MISMATCH_FORCE;MSG_U200PTI_UPDATE_PTI_FAILED_ABORT;MSG_U200PTI_UPDATE_PTI_FAILED_NOTRAIN;MSG_U200PTI_UPDATE_PTI_FAILED_RETRY;MSG_U200PTI_UPDATE_PTI_OK;MSG_U200PTI_UPDATE_ROLLINGSTOCKID;MSG_U200PTI_UPDATE_ROLLINGSTOCKID_MISMATCH;MSG_U200PTI_UPDATE_ROLLINGSTOCKID_MISMATCH_FORCE;MSG_U200PTI_UPDATE_TRAINID;MSG_U200PTI_UPDATE_TRAINID_MISMATCH;INTERLOCKINGCTRL_DESTNIATION_NOT_ACCEPTABLE;MSG_511_TRAINMOVEINNONSIGNALEDAREA;MSG_511_TRAINMOVEINSIGNALEDAREA;MSG_553_TRAINSWAPINNONSIGNALEDAREA;MSG_TDS_152_POSSIBLETRAINMERGE;MSG_TDS_153_UNSCHEDULEDTRAINSPLITTING;MSG_TDS_154_ENDOFFALSEOCCUPANCY;MSG_TDS_154_FALSEOCCUPANCY;MSG_TDS_155_DISAPPEAREDTRAINDELETION;MSG_TDS_155_NSDISAPPEAREDTRAINDELETION;MSG_TDS_156_RECOVER_DISAPPEAREDTRAIN;MSG_TDS_157_TRAINCHANGESDIRECTION;MSG_TDS_158_FAILED_TO_RESET_TC;MSG_TDS_159_FAILED_TO_SHUNT_TC;MSG_TDS_159_FALSERELEASE_TC;MSG_TDS_161_TRAINENTERTC;MSG_TDS_161_TRAINLEAVETC;MSG_TDS_164_TRAIN_ENTRANCE;MSG_TDS_164_TRAIN_EXIT;MSG_TDS_175_DRIVERID;MSG_TDS_175_ROLLINGSTOCKID;MSG_TDS_176_DRIVERID;MSG_TDS_176_RENUMBERING;MSG_TDS_176_ROLLINGSTOCKID;MSG_TDS_178_MAXTRAINBYTC;MSG_TDS_179_PARAMETERIDCHANGED;MSG_TDS_179_PARAMETERIDSET;MSG_TDS_180_PROMOTE_TRAIN2FO;MSG_TDS_197_FO_TURNED_TO_TRAIN;MSG_TDS_198_CREATE_TRAINONTC;MSG_TDS_302_NOTIDENTIFICATIONWITHINDELAY;MSG_TDS_303_INSERVICE;MSG_TDS_303_OUTOFSERVICE;MSG_TDS_307_DECLASSIFY_TRAIN;MSG_TDS_308_PROBABLE_FALSEOCCUPANCY;MSG_TDS_309_END_PROBABLE_FALSEOCCUPANCY;MSG_TDS_310_POSSIBLEDISAPPEAREDTRAIN;MSG_TDS_322_TRAINOVERRIDENFO;MSG_TDS_418_MANUAL_TRAIN_MERGING;MSG_TDS_419_MERGE_DUE_TO_TC_LIMITATION;MSG_TDS_426_HOTBOX;MSG_TDS_431_TRAINCHANGESDIRECTION;MSG_TDS_444_ATDAPPROACHINGTRAIN;MSG_TDS_444_ATDCLEAROUTTRAIN;MSG_TDS_444_ATDTRAINTIMEOUT;MSG_TDS_444_ATDUPDATINGTRAIN;MSG_TDS_451_WAKEUP;MSG_TDS_452_IDLE;MSG_TDS_505_NOMORETCOCCUPANCYTOOLONG;MSG_TDS_505_TCOCCUPANCYTOOLONG;MSG_TDS_513_MANUALSPLIT;MSG_TDS_516_CONTROLLEDAREA;MSG_TDS_517_TRAIN_ENTRY_NONSIGNALED;MSG_TDS_517_TRAIN_EXIT_NONSIGNALED;MSG_TDS_554_APPROACHTRIPPER;MSG_TDS_PROMOTE_TC2FO;MSG_TDS_STATEDETECTOR_NOTRAIN;ROUTESETTING_COMMANDROUTE_ANSWER_NOK;ROUTESETTING_COMMANDROUTE_ANSWER_OK;ROUTESETTING_COMMANDROUTE_REQUESTED;ROUTESETTING_INTERPOSEDTRAIN;ROUTESETTING_MODIFY_MANEUVER_OF_TRIP;ROUTESETTING_MODIFY_POINT_OF_TRIP;ROUTESETTING_NOTPERMITTEDTOSETAROUTE;ROUTESETTING_TRAINISSTUCK;ROUTESETTING_UNABLETOSETALTERNATIVEROUTE;ROUTESETTING_UNABLETOSETAROUTE;ROUTESETTING_USERDIALOG_ACCEPTED;ROUTESETTING_USERDIALOG_REFUSED;ROUTESETTING_USERDIALOG_TIMEOUT;ROUTESETTING_USERDIALOG_UNKNOWN</Property>\n')
        f.write('            <Property dt="i4" name="MaxFailedPerHour">0</Property>\n')
        f.write('            <Property dt="boolean" name="TrackingMode">false</Property>\n')
        f.write('            <Property dt="i4" name="UEID">'+str(u)+'</Property>\n')
        f.write('            <MultiLingualProperty name="Name">\n')
        f.write('              <MultiLingualValue localeId="1033" roleId="-1">TC'+check(str(sh.cell(row=i,column=2).value))+'</MultiLingualValue>\n')
        f.write('              <MultiLingualValue localeId="1036" roleId="-1">TC'+check(str(sh.cell(row=i,column=2).value))+'</MultiLingualValue>\n')
        f.write('            </MultiLingualProperty>\n')
        f.write('''            <PropertyList name="RoutesForInhibition">
              <ListElem dt="string" index="1">R</ListElem>
            </PropertyList>\n''')
        f.write('          </Properties>\n')
        f.write('        </Object>\n')
        i+=1
        u+=1
    f.write('      </Objects>\n')
    f.write('    </Class>\n')
    f.write('  </Classes>\n')
    f.write('</ObjectPropertyModule>')


