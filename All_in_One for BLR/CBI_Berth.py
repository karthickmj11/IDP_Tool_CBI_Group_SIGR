
import openpyxl
def run(loc,op_l):
    wb = openpyxl.load_workbook(loc)
    f = open(op_l+'\\Berth.xml','w')
    sh = wb['Berth']
    i=2
    f.write('<Classes>\n')
    f.write('<Class name="Berth" rules="update" traces="error">\n')
    f.write('<Objects>\n')
    while(sh.cell(row=i,column=2).value!=None):
        f.write('<Object name="'+str(sh.cell(row=i,column=2).value)+'" rules="update_or_create">\n')
        f.write('<Properties>\n')
        f.write('<PropertyList name="InterStation">\n')
        f.write('</PropertyList>\n')
        f.write('<Property name="ID" dt="string">'+str(sh.cell(row=i,column=2).value)+'</Property>\n')
        f.write('<MultiLingualProperty name="Name">\n')
        f.write('<MultiLingualValue roleId="-1" localeId="1033">'+str(sh.cell(row=i,column=2).value)+'</MultiLingualValue>\n')
        f.write('<MultiLingualValue roleId="-1" localeId="1036">'+str(sh.cell(row=i,column=2).value)+'</MultiLingualValue>\n')
        f.write('</MultiLingualProperty>\n')
        f.write('</Properties>\n')
        f.write('</Object>\n')
        i+=1
    f.write('</Objects>\n')
    f.write('</Class>\n')
    f.write('</Classes>\n')

