import openpyxl
def run(loc,op_l):
    wb = openpyxl.load_workbook(loc)
    f = open(op_l+'\\Group_PhysicalPoint HMI.xml','w')
    sh = wb['PSAPR']
    sh1 = wb['Interlocking']
    i=2
    u = str(sh1.cell(row=1,column=1).value)
    f.write('<Classes>\n')
    f.write('<Class name="PhysicalPointHMI" rules="update" traces="error">\n')
    f.write('<Objects>\n')
    f.write('        <Object name="Stop_STA_'+str(sh1.cell(row=1,column=1).value)+'_D" rules="update" traces="error">\n')
    f.write('          <Properties>\n')
    f.write('            <Property dt="string" name="AreaGroup">Area/LI_1/MI_1/Territory_'+str(u)+'/Area_'+str(u)+'</Property>\n')
    f.write('            <Property dt="string" name="FunctionGroup">Function/Signalling/PhysicalPointHMI/Stop_STA_'+str(sh1.cell(row=1,column=1).value)+'_D</Property>\n')
    f.write('          </Properties>\n')
    f.write('        </Object>\n')
    f.write('        <Object name="Stop_STA_'+str(sh1.cell(row=1,column=1).value)+'_U" rules="update" traces="error">\n')
    f.write('          <Properties>\n')
    f.write('            <Property dt="string" name="AreaGroup">Area/LI_1/MI_1/Territory_'+str(u)+'/Area_'+str(u)+'</Property>\n')
    f.write('            <Property dt="string" name="FunctionGroup">Function/Signalling/PhysicalPointHMI/Stop_STA_'+str(sh1.cell(row=1,column=1).value)+'_U</Property>\n')
    f.write('          </Properties>\n')
    f.write('        </Object>\n')
    while(sh.cell(row=i,column=2).value!=None):
        f.write('        <Object name="Stop_STA_'+str(sh.cell(row=1,column=2).value)+'_D" rules="update" traces="error">\n')
        f.write('          <Properties>\n')
        f.write('            <Property dt="string" name="AreaGroup">Area/LI_1/MI_1/Territory_'+str(u)+'/Area_'+str(u)+'</Property>\n')
        f.write('            <Property dt="string" name="FunctionGroup">Function/Signalling/PhysicalPointHMI/Stop_STA_'+str(sh.cell(row=i,column=2).value)+'_D</Property>\n')
        f.write('          </Properties>\n')
        f.write('        </Object>\n')
        f.write('        <Object name="Stop_STA_'+str(sh.cell(row=1,column=2).value)+'_U" rules="update" traces="error">\n')
        f.write('          <Properties>\n')
        f.write('            <Property dt="string" name="AreaGroup">Area/LI_1/MI_1/Territory_'+str(u)+'/Area_'+str(u)+'</Property>\n')
        f.write('            <Property dt="string" name="FunctionGroup">Function/Signalling/PhysicalPointHMI/Stop_STA_'+str(sh.cell(row=i,column=2).value)+'_U</Property>\n')
        f.write('          </Properties>\n')
        f.write('        </Object>\n')
        i+=1
    f.write('\n')
    f.write('\n')
    f.write('\n')





