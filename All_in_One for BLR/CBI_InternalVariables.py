#input for this file is total signal object present in this sector
import openpyxl
def run(loc,op_l):
    wb = openpyxl.load_workbook(loc)
    sh = wb['Signal']
    i=2
    f = open(op_l+'\\InternalVariables.xml','w')
    s=''
    t=''
    l=''
    f.write('    <Class name="Variable">\n')
    f.write('      <Objects>\n')
    while(sh.cell(row=i,column=2).value!=None):
        c = str(sh.cell(row=i,column=2).value)[1:]
        s+='        <Object name="S_REL'+c+'" rules="update_or_create">\n'
        s+='          <Properties>\n'
        s+='            <Property dt="string" name="Type">LM</Property>\n'
        s+='          </Properties>\n'
        s+='        </Object>\n'
        t+='        <Object name="S_SEL'+c+'" rules="update_or_create">\n'
        t+='          <Properties>\n'
        t+='            <Property dt="string" name="Type">LM</Property>\n'
        t+='          </Properties>\n'
        t+='        </Object>\n'
        l+='        <Object name="S_SET'+c+'" rules="update_or_create">\n'
        l+='          <Properties>\n'
        l+='            <Property dt="string" name="Type">LM</Property>\n'
        l+='          </Properties>\n'
        l+='        </Object>\n'
        i+=1
    f.write(s)
    f.write(t)
    f.write(l)
    f.write('      </Objects>\n')
    f.write('    </Class>\n')
