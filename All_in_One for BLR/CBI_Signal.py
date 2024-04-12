import openpyxl
def run(loc,op_l):
    wb = openpyxl.load_workbook(loc)
    f = open(op_l+'\\Signal.xml','w')
    #f.write('<!-- !!!Please write SPAD manually!!! -->\n')
    f.write('''<?xml version="1.0" ?>
<ObjectPropertyModule Generation_Date="2019-02-25" Project="BLR" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="Signal.xsd">
  <Versions>
    <Version Name="ATS_SYSTEM_PARAMETERS#Comment" Value="XML file containing the System Data"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Date" Value="2011-03-19"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#SyPD_Version" Value="_none_"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#UEVOLtoIDP_Version" Value="2.1.7"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Writer_Name" Value="IconisDigesterCore 0.0.4091.29806"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#XML_File_Version" Value="1.6.6.1078"/>
    <Version Name="ICONIS_ATS_Equipment#SyID_Version" Value="Y3-64 A427945-J1"/>
    <Version Name="ICONIS_IDP#Customisation_SwRSAD_Version" Value="none"/>
    <Version Name="ICONIS_IDP#Product_SwRSAD_Version" Value="Y3-64 A427875-A"/>
    <Version Name="ICONIS_IDP#Product_Version" Value="10.3.4 (revision 3010)"/>
    <Version Name="ICONIS_IDP#Project_Version" Value="BLR#V10.3.4 (revision 3060)"/>
    <Version Name="Project#ATS_Custom_Reference_Database_Version" Value="0.0.3.565"/>
    <Version Name="Project#Database_Version" Value="0.0.3.565"/>
  </Versions>\n''')
    f.write('  <Classes>\n')
    f.write('    <Class name="Signal">\n')
    f.write('      <Objects>\n')
    sh = wb['Signal']
    sh1= wb['Route']
    i=2
    sector = str(sh.cell(row=i,column=2).value)[-1]
    while(sh.cell(row=i,column=2).value!=None):
        c1 = str(sh.cell(row=i,column=2).value)[1:]
        f.write('        <Object name="S'+c1+'" rules="update_or_create">\n')
        f.write('          <Properties>\n')
        f.write('            <Property dt="boolean" name="ApproachLockingAvailable">true</Property>\n')
        f.write('            <Property dt="boolean" name="HILCAvailable">true</Property>\n')
        f.write('            <Property dt="string" name="Interlocking">ASCV_'+sector+'</Property>\n')
        f.write('            <Property dt="string" name="ManagedKeys">APPROACH_SG_NOTSET;APPROACH_SG_SET;APPROACH_SG_SET_RD;S_BLUE;S_FAILURE;S_GREEN;S_NOTPROVED_G;S_NOTPROVED_R;S_RED;S_ROUTE_INDICATOR_NOTSET;S_ROUTE_INDICATOR_SET;SI_CONF_SBL_RES;SI_CONF_SUBL_RES;SI_PREP_SBL_RES;SI_PREP_SUBL_RES</Property>\n')
        if(sh.cell(row=i,column=33).value!=None):
            f.write('            <Property dt="string" name="SPAD">'+str(sh.cell(row=i,column=33).value)+'</Property>\n')
        else:
            f.write('            <Property dt="string" name="SPAD"></Property>\n')
            
        f.write('            <Property dt="boolean" name="OverriddenAvailable">true</Property>\n')
        f.write('            <Property dt="boolean" name="CallOnSignalStatusAvailable">false</Property>\n')
        f.write('            <Property dt="string" name="EnableInstanceSPAD">true</Property>\n')
        f.write('            <Property dt="i4" name="HILCDestinationBlockingType">0</Property>\n')
        f.write('            <Property dt="i4" name="HILCDestinationBlockingWithConfirmation">1</Property>\n')
        f.write('            <Property dt="i4" name="HILCWithConfirmation">1</Property>\n')
        f.write('            <Property dt="i4" name="LampProvedTypeAvailable">0</Property>\n')
        f.write('            <MultiLingualProperty name="DestinationSignals">\n')
        s=''
        j=2
        while(sh1.cell(row=j,column=2).value!=None):
            l = str(sh1.cell(row=j,column=2).value)[1:].split('_')
            if(c1[-4:] == l[0]):
                s+='Signal Name=&quot;S'+l[1]+'&quot; ID=&quot;S'+l[1]+'&quot; RouteID=&quot;R'+str(sh1.cell(row=j,column=2).value)[1:]+'&quot; CallOn=&quot;0&quot; Auto=&quot;0&quot; OppositeName=&quot;&quot; OppositeID=&quot;0&quot;/&gt;&lt;'
            j+=1
        f.write('              <MultiLingualValue localeId="1033" roleId="-1">&lt;Signals&gt;&lt;'+s+'/Signals&gt;</MultiLingualValue>\n')
        f.write('              <MultiLingualValue localeId="1036" roleId="-1">&lt;Signals&gt;&lt;'+s+'/Signals&gt;</MultiLingualValue>\n')
        f.write('            </MultiLingualProperty>\n')
        #f.write('            <MultiLingualProperty name="OriginSignals">\n')
        #s=''
        #j=2
        #while(sh1.cell(row=j,column=2).value!=None):
        #    l = str(sh1.cell(row=j,column=2).value)[1:].split('_')
        #    if(c1[-4:] == l[1]):
        #        s+='Signal Name=&quot;S'+l[0]+'&quot; ID=&quot;S'+l[0]+'&quot; RouteID=&quot;R'+str(sh1.cell(row=j,column=2).value)[1:]+'&quot; CallOn=&quot;0&quot; Auto=&quot;0&quot; OppositeName=&quot;&quot; OppositeID=&quot;0&quot;/&gt;&lt;'
        #    j+=1
        #f.write('              <MultiLingualValue localeId="1033" roleId="-1">&lt;Signals&gt;&lt;'+s+'/Signals&gt;</MultiLingualValue>\n')
        #f.write('              <MultiLingualValue localeId="1036" roleId="-1">&lt;Signals&gt;&lt;'+s+'/Signals&gt;</MultiLingualValue>\n')
        #f.write('            </MultiLingualProperty>\n')
        f.write('            <MultiLingualProperty name="Name">\n')
        f.write('              <MultiLingualValue localeId="1033" roleId="-1">S'+c1+'</MultiLingualValue>\n')
        f.write('              <MultiLingualValue localeId="1036" roleId="-1">S'+c1+'</MultiLingualValue>\n')
        f.write('            </MultiLingualProperty>\n')
        f.write('          </Properties>\n')
        f.write('        </Object>\n')
        i+=1
    f.write('      </Objects>\n')
    f.write('    </Class>\n')
    f.write('  </Classes>\n')
    f.write('</ObjectPropertyModule>\n')
