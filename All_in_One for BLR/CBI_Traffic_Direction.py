import openpyxl
def run(loc,op_l):
    wb = openpyxl.load_workbook(loc)
    sh = wb['TD']
    f = open(op_l+'\\TrafficDirection.xml','w')
    f.write('''<?xml version="1.0" ?>
<ObjectPropertyModule Generation_Date="2019-02-25" Project="BLR" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="TrafficDirection.xsd">
  <Versions>
    <Version Name="ATS_SYSTEM_PARAMETERS#Comment" Value="XML file containing the System Data"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Date" Value="2011-03-19"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#SyPD_Version" Value="_none_"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#UEVOLtoIDP_Version" Value="2.1.7"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Writer_Name" Value="IconisDigesterCore 0.0.4091.29806"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#XML_File_Version" Value="1.6.6.1078"/>
    <Version Name="ICONIS_ATS_Equipment#SyID_Version" Value="Y3-64 A427945-J1"/>
    <Version Name="ICONIS_IDP#Customisation_SwRSAD_Version" Value="none"/>
    <Version Name="ICONIS_IDP#Product_SwRSAD_Version" Value="Y3-64 A427875-A"/>
    <Version Name="ICONIS_IDP#Product_Version" Value="10.3.4 (revision 3010)"/>
    <Version Name="ICONIS_IDP#Project_Version" Value="BLR#V10.3.4 (revision 3060)"/>
    <Version Name="Project#ATS_Custom_Reference_Database_Version" Value="0.0.3.565"/>
    <Version Name="Project#Database_Version" Value="0.0.3.565"/>
  </Versions>
\n''')
    f.write('  <Classes>\n')
    f.write('    <Class name="TrafficDirection">\n')
    f.write('      <Objects>\n')
    i=2
    l = []
    while(sh.cell(row=i,column=2).value!=None):
        c = str(sh.cell(row=i,column=2).value)[3:].split('_')
        if(len(c)==2):
            if c[0] not in l:
                #print(c[0])
                l.append(c[0])
                f.write('        <Object name="TD_'+c[0]+'" rules="update_or_create">\n')
                f.write('          <Properties>\n')
                f.write('            <MultiLingualProperty name="Name">\n')
                f.write('              <MultiLingualValue localeId="1033" roleId="-1">TD_'+c[0]+'</MultiLingualValue>\n')
                f.write('              <MultiLingualValue localeId="1036" roleId="-1">TD_'+c[0]+'</MultiLingualValue>\n')
                f.write('            </MultiLingualProperty>\n')
                f.write('          </Properties>\n')
                f.write('        </Object>\n')
        else:
            if c[0]+'_'+c[1] not in l:
                l.append(c[0]+'_'+c[1])
                f.write('        <Object name="TD_'+c[0]+'_'+c[1]+'" rules="update_or_create">\n')
                f.write('          <Properties>\n')
                f.write('            <MultiLingualProperty name="Name">\n')
                f.write('              <MultiLingualValue localeId="1033" roleId="-1">TD_'+c[0]+'_'+c[1]+'</MultiLingualValue>\n')
                f.write('              <MultiLingualValue localeId="1036" roleId="-1">TD_'+c[0]+'_'+c[1]+'</MultiLingualValue>\n')
                f.write('            </MultiLingualProperty>\n')
                f.write('          </Properties>\n')
                f.write('        </Object>\n')
        f.write('        <Object name="'+str(sh.cell(row=i,column=2).value)+'" rules="update_or_create">\n')
        f.write('          <Properties>\n')
        f.write('            <MultiLingualProperty name="Name">\n')
        f.write('              <MultiLingualValue localeId="1033" roleId="-1">'+str(sh.cell(row=i,column=2).value)+'</MultiLingualValue>\n')
        f.write('              <MultiLingualValue localeId="1036" roleId="-1">'+str(sh.cell(row=i,column=2).value)+'</MultiLingualValue>\n')
        f.write('            </MultiLingualProperty>\n')
        f.write('          </Properties>\n')
        f.write('        </Object>\n')
        i+=1
    f.write('      </Objects>\n')
    f.write('    </Class>\n')
    f.write('  </Classes>\n')
    f.write('</ObjectPropertyModule>')


