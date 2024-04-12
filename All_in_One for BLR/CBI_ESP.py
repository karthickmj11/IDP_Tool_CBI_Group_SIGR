import openpyxl
def run(loc,op_l):
    wb = openpyxl.load_workbook(loc)
    f = open(op_l+'\\ESP.xml','w')
    sh = wb['ESP']
    i=2
    f.write('''<?xml version="1.0" ?>
<ObjectPropertyModule Generation_Date="2019-02-25" Project="BLR" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="ESP.xsd">
  <Versions>
    <Version Name="ATS_SYSTEM_PARAMETERS#Comment" Value="XML file containing the System Data"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Date" Value="2011-03-19"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#SyPD_Version" Value="_none_"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#UEVOLtoIDP_Version" Value="2.1.7"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Writer_Name" Value="IconisDigesterCore 0.0.4091.29806"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#XML_File_Version" Value="1.6.6.1078"/>
    <Version Name="ICONIS_ATS_Equipment#SyID_Version" Value="Y3-64 A427945-J1"/>
    <Version Name="ICONIS_IDP#Customisation_SwRSAD_Version" Value="none"/>
    <Version Name="ICONIS_IDP#Product_SwRSAD_Version" Value="Y3-64 A427875-A"/>
    <Version Name="ICONIS_IDP#Product_Version" Value="10.3.4 (revision 3010)"/>
    <Version Name="ICONIS_IDP#Project_Version" Value="BLR#V10.3.4 (revision 3060)"/>
    <Version Name="Project#ATS_Custom_Reference_Database_Version" Value="0.0.3.565"/>
    <Version Name="Project#Database_Version" Value="0.0.3.565"/>
  </Versions>\n''')
    f.write('<Classes>\n')
    f.write('<Class name="ESP">\n')
    f.write('<Objects>\n')
    while(sh.cell(row=i,column=2).value!=None):
        f.write('        <Object name="ESP_'+str(sh.cell(row=i,column=2).value)[3:]+'" rules="update_or_create">\n')
        f.write('          <Properties>\n')
        f.write('            <Property dt="string" name="Status_PV">'+str(sh.cell(row=i,column=2).value)+'</Property>\n')
        f.write('          </Properties>\n')
        f.write('        </Object>\n')
        i+=1
    f.write('      </Objects>\n')
    f.write('</Class>\n')
    f.write('</Classes>\n')
    f.write('</ObjectPropertyModule>')





