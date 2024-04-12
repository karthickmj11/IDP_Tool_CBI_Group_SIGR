import openpyxl
def run(loc,op_l):
    wb = openpyxl.load_workbook(loc)
    f = open(op_l+'\\Group_Traffic_Direction.xml','w')
    sh = wb['TD']
    sh1 = wb['Interlocking']
    i=2
    u = str(sh1.cell(row=1,column=1).value)
    f.write('  <Classes>\n')
    f.write('    <Class name="TrafficDirection" rules="update" traces="error">\n')
    f.write('      <Objects>\n')
    while(sh.cell(row=i,column=2).value!=None):
        f.write('        <Object name="'+str(sh.cell(row=i,column=2).value)+'" rules="update" traces="error">\n')
        f.write('          <Properties>\n')
        f.write('            <Property dt="string" name="AreaGroup">Area/LI_1/MI_1/Territory_'+str(u)+'/Area_'+str(u)+'</Property>\n')
        f.write('            <Property dt="string" name="FunctionGroup">Function/Signalling/Traffic_Direction/'+str(sh.cell(row=i,column=2).value)+'</Property>\n')
        f.write('          </Properties>\n')
        f.write('        </Object>\n')
        i+=1
    f.write('      </Objects>\n')
    f.write('    </Class>\n')
    f.write('  </Classes>\n')





