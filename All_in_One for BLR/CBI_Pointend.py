import openpyxl
def run(loc,op_l):
    wb = openpyxl.load_workbook(loc)
    f = open(op_l+'\\Pointend.xml','w')
    sh = wb['Point']
    sh1 = wb['Switch']
    i=2
    def ch():
        j=1
        s=''
        while(sh1.cell(row=j,column=2).value!=None):
            if(len(str(sh1.cell(row=j,column=2).value)[2:])==7):
                l=str(sh1.cell(row=j,column=2).value)[2:].split('_')
                if(l[0]==str(sh.cell(row=i,column=2).value)[1:] ):
                    s += str(sh1.cell(row=j,column=2).value)[2:]
                elif(l[1]==str(sh.cell(row=i,column=2).value)[1:]):
                    s += str(sh1.cell(row=j,column=2).value)[2:]
                
            else:
                s+=''
            j+=1
        return s
    f.write('''<?xml version="1.0" ?>
<ObjectPropertyModule Generation_Date="2019-02-25" Project="BLR" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="PointEnd.xsd">
  <Versions>
    <Version Name="ATS_SYSTEM_PARAMETERS#Comment" Value="XML file containing the System Data"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Date" Value="2011-03-19"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#SyPD_Version" Value="_none_"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#UEVOLtoIDP_Version" Value="2.1.7"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Writer_Name" Value="IconisDigesterCore 0.0.4091.29806"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#XML_File_Version" Value="1.6.6.1078"/>
    <Version Name="ICONIS_ATS_Equipment#SyID_Version" Value="Y3-64 A427945-J1"/>
    <Version Name="ICONIS_IDP#Customisation_SwRSAD_Version" Value="none"/>
    <Version Name="ICONIS_IDP#Product_SwRSAD_Version" Value="Y3-64 A427875-A"/>
    <Version Name="ICONIS_IDP#Product_Version" Value="10.3.4 (revision 3010)"/>
    <Version Name="ICONIS_IDP#Project_Version" Value="BLR#V10.3.4 (revision 3060)"/>
    <Version Name="Project#ATS_Custom_Reference_Database_Version" Value="0.0.3.565"/>
    <Version Name="Project#Database_Version" Value="0.0.3.565"/>
  </Versions>
\n''')
    f.write('  <Classes>\n')
    f.write('    <Class name="PointEnd">\n')
    f.write('      <Objects>\n')
    while(sh.cell(row=i,column=2).value!=None):
        s=''
        t=''
        j=2
        while(sh1.cell(row=j,column=2).value!=None):
            if(len(str(sh1.cell(row=j,column=2).value)[2:])>=7):
                l=str(sh1.cell(row=j,column=2).value)[2:].split('_')
                if(l[0]==str(sh.cell(row=i,column=2).value)[1:] ):
                    s += str(sh1.cell(row=j,column=2).value)[2:]
                    t += l[1]
                
                elif(l[1]==str(sh.cell(row=i,column=2).value)[1:]):
                    s += str(sh1.cell(row=j,column=2).value)[2:]
                    t += l[0]
                
            else:
                s+=''
                t+=''
            j+=1
        f.write('        <Object name="'+str(sh.cell(row=i,column=2).value)+'" rules="update_or_create">\n')
        f.write('          <Properties>\n')
        f.write('            <Property dt="boolean" name="HILCAvailable">true</Property>\n')
        f.write('            <Property dt="i4" name="HILCWithConfirmation">1</Property>\n')
        f.write('            <Property dt="string" name="ManagedKeys">PT_OUTOF_CORRESPONDENCE</Property>\n')
        f.write('            <Property dt="string" name="SwitchID">SW'+s+'</Property>\n')
        f.write('            <MultiLingualProperty name="ImpactedPointEnds">\n')
        

        f.write('              <MultiLingualValue localeId="1033" roleId="-1">&lt;PointEnds Switch=&quot;SW'+s+'&quot;&gt;&lt;PointEnd Name=&quot;P'+t+'&quot; ID=&quot;P'+t+'&quot; Switch=&quot;SW'+s+'&quot;/&gt;&lt;/PointEnds&gt;</MultiLingualValue>\n')
        f.write('              <MultiLingualValue localeId="1036" roleId="-1">&lt;PointEnds Switch=&quot;SW'+s+'&quot;&gt;&lt;PointEnd Name=&quot;P'+t+'&quot; ID=&quot;P'+t+'&quot; Switch=&quot;SW'+s+'&quot;/&gt;&lt;/PointEnds&gt;</MultiLingualValue>\n')
        f.write('            </MultiLingualProperty>\n')
        f.write('            <MultiLingualProperty name="Name">\n')
        f.write('              <MultiLingualValue localeId="1033" roleId="-1">'+str(sh.cell(row=i,column=2).value)+'</MultiLingualValue>\n')
        f.write('              <MultiLingualValue localeId="1036" roleId="-1">'+str(sh.cell(row=i,column=2).value)+'</MultiLingualValue>\n')
        f.write('            </MultiLingualProperty>\n')
        f.write('            <MultiLingualProperty name="SwitchName">\n')
        f.write('              <MultiLingualValue localeId="1033" roleId="-1">SW'+s+'</MultiLingualValue>\n')
        f.write('              <MultiLingualValue localeId="1036" roleId="-1">SW'+s+'</MultiLingualValue>\n')
        f.write('            </MultiLingualProperty>\n')
        f.write('          </Properties>\n')
        f.write('        </Object>\n')
        i+=1
    f.write('      </Objects>\n')
    f.write('    </Class>\n')
    f.write('  </Classes>\n')
    f.write('</ObjectPropertyModule>')


