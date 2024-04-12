import openpyxl
def dir(a):
        if(int(a[0:2])%2==0):
            return "LEFT"
        else:
            return "RIGHT"
def run(loc,op_l):
    wb = openpyxl.load_workbook(loc)
    f = open(op_l+'\\TrackPortion.xml','w')
    sh = wb['TP']
    sh1 = wb['Interlocking']
    f.write('''<?xml version="1.0" ?>
<ObjectPropertyModule Generation_Date="2019-02-25" Project="BLR" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="TrackPortion.xsd">
  <Versions>
    <Version Name="ATS_SYSTEM_PARAMETERS#Comment" Value="XML file containing the System Data"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Date" Value="2011-03-19"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#SyPD_Version" Value="_none_"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#UEVOLtoIDP_Version" Value="2.1.7"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#Writer_Name" Value="IconisDigesterCore 0.0.4091.29806"/>
    <Version Name="ATS_SYSTEM_PARAMETERS#XML_File_Version" Value="1.6.6.1078"/>
    <Version Name="ICONIS_ATS_Equipment#SyID_Version" Value="Y3-64 A427945-J1"/>
    <Version Name="ICONIS_IDP#Customisation_SwRSAD_Version" Value="none"/>
    <Version Name="ICONIS_IDP#Product_SwRSAD_Version" Value="Y3-64 A427875-A"/>
    <Version Name="ICONIS_IDP#Product_Version" Value="10.3.4 (revision 3010)"/>
    <Version Name="ICONIS_IDP#Project_Version" Value="BLR#V10.3.4 (revision 3060)"/>
    <Version Name="Project#ATS_Custom_Reference_Database_Version" Value="0.0.3.565"/>
    <Version Name="Project#Database_Version" Value="0.0.3.565"/>
  </Versions>
\n''')
    f.write('  <Classes>\n')
    f.write('    <Class name="TrackPortion">\n')
    f.write('      <Objects>\n')
    i=2
    u = str(sh1.cell(row=1,column=1).value)
    sector = str(sh1.cell(row=1,column=2).value)
    while(sh.cell(row=i,column=2).value!=None):
        f.write('        <Object name="'+str(sh.cell(row=i,column=2).value)+'" rules="update_or_create">\n')
        f.write('          <Properties>\n')
        f.write('            <Property dt="i4" name="KPBegin"></Property>\n')
        f.write('            <Property dt="string" name="Orientation">'+dir(str(sh.cell(row=i,column=2).value)[2:])+'</Property>\n')
        f.write('            <Property dt="i4" name="SDDCheckableStatus">0</Property>\n')
        f.write('            <Property dt="i4" name="SDDClearableStatus">0</Property>\n')
        f.write('            <Property dt="i4" name="SDDTypeStatus">0</Property>\n')
        f.write('            <Property dt="string" name="TrackerOperationalStatus">T'+str(sh.cell(row=i,column=2).value)[2:5]+'</Property>\n')
        s='/>'
        q='/>'
        if(sh.cell(row=i,column=1).value!=None):
            s='>&lt;FixedBlocks&gt;&lt;FixedBlock Name=&quot;TC'+str(sh.cell(row=i,column=1).value)[2:4]+sector+'&quot; ID=&quot;'+str(sh.cell(row=i,column=1).value)+'&quot; CBI=&quot;ASCV_'+sector+'&quot; Area=&quot;Area_'+u+'&quot;/&gt;&lt;/FixedBlocks&gt;</MultiLingualValue>'

        f.write('            <MultiLingualProperty name="LeftBlocks">\n')
        f.write('              <MultiLingualValue localeId="1033" roleId="-1"'+s+'\n')
        f.write('              <MultiLingualValue localeId="1036" roleId="-1"'+s+'\n')
        f.write('           </MultiLingualProperty>\n')
        f.write('            <MultiLingualProperty name="RightBlocks">\n')
        if(sh.cell(row=i,column=3).value!=None):
            q='>&lt;FixedBlocks&gt;&lt;FixedBlock Name=&quot;TC'+str(sh.cell(row=i,column=3).value)[2:4]+sector+'&quot; ID=&quot;'+str(sh.cell(row=i,column=3).value)+'&quot; CBI=&quot;ASCV_'+sector+'&quot; Area=&quot;Area_'+u+'&quot;/&gt;&lt;/FixedBlocks&gt;</MultiLingualValue>'
        f.write('              <MultiLingualValue localeId="1033" roleId="-1"'+q+'\n')
        f.write('              <MultiLingualValue localeId="1036" roleId="-1"'+q+'\n')
        f.write('            </MultiLingualProperty>\n')
        f.write('            <MultiLingualProperty name="Name">\n')
        f.write('              <MultiLingualValue localeId="1033" roleId="-1">TC'+str(sh.cell(row=i,column=2).value)[2:4]+sector+'</MultiLingualValue>\n')
        f.write('              <MultiLingualValue localeId="1036" roleId="-1">TC'+str(sh.cell(row=i,column=2).value)[2:4]+sector+'</MultiLingualValue>\n')
        f.write('            </MultiLingualProperty>\n')
        f.write('          </Properties>\n')
        f.write('        </Object>\n')
        i+=1
    f.write('      </Objects>\n')
    f.write('    </Class>\n')
    f.write('  </Classes>\n')
    f.write('</ObjectPropertyModule>')

