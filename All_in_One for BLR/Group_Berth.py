import openpyxl
def run(loc,op_l):
    wb = openpyxl.load_workbook(loc)
    f = open(op_l+'\\Group_Berth.xml','w')
    sh = wb['Berth']
    sh1 = wb['Interlocking']
    i=2
    u = sh1.cell(row=1,column=1).value
    f.write('''<?xml version="1.0"?>
<ObjectPropertyModule name="AreaFunctionGroup" Project="BLR" Generation_Date="2011-03-20" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="OPFormat.xsd"  >
<Versions>
<Version Name="ATS_SYSTEM_PARAMETERS#Comment" Value="XML file containing the System Data"/>
<Version Name="ATS_SYSTEM_PARAMETERS#Date" Value="2011-03-19"/>
<Version Name="ATS_SYSTEM_PARAMETERS#SyPD_Version" Value="_none_"/>
<Version Name="ATS_SYSTEM_PARAMETERS#UEVOLtoIDP_Version" Value="2.1.5"/>
<Version Name="ATS_SYSTEM_PARAMETERS#Writer_Name" Value="IconisDigesterCore 0.0.4091.29806"/>
<Version Name="ATS_SYSTEM_PARAMETERS#XML_File_Version" Value="1.6.6.1078"/>
<Version Name="ICONIS_ATS_Equipment#SyID_Version" Value="Y3-64 A427945-D9"/>
<Version Name="ICONIS_IDP#Customisation_SwRSAD_Version" Value="none"/>
<Version Name="ICONIS_IDP#Product_SwRSAD_Version" Value="Y3-64 A427875-A"/>
<Version Name="ICONIS_IDP#Product_Version" Value="3.10.1.1060"/>
<Version Name="ICONIS_IDP#Project_Version" Value="BLR#V1.1.0.1070"/>
<Version Name="Project#ATS_Custom_Reference_Database_Version" Value="0.0.3.565"/>
<Version Name="Project#Database_Version" Value="0.0.3.565"/>
</Versions>\n''')
    f.write('<Classes>\n')
    f.write('<Class name="Berth" rules="update" traces="error">\n')
    f.write('<Objects>\n')
    while(sh.cell(row=i,column=2).value!=None):
        f.write('        <Object name="'+str(sh.cell(row=i,column=2).value)+'" rules="update" traces="error">\n')
        f.write('          <Properties>\n')
        f.write('            <Property name="AreaGroup" dt="string">Area/LI_1/MI_1/Territory_'+str(u)+'/Area_'+str(u)+'</Property>\n')
        f.write('            <Property name="FunctionGroup" dt="string">Function/Signalling/Berth/'+str(sh.cell(row=i,column=2).value)+'</Property>\n')
        f.write('          </Properties>\n')
        f.write('        </Object>\n')
        i+=1
    f.write('      </Objects>\n')
    f.write('</Class>\n')
    f.write('</Classes>\n')
    f.write('</ObjectPropertyModule>')



