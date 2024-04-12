
import openpyxl
def run(loc,op_l):
    wb = openpyxl.load_workbook(loc)
    f = open(op_l+'\\Group_Junction.xml','w')
    sh1 = wb['Interlocking']
    i=2
    u = sh1.cell(row=1,column=1).value
    f.write('  <Classes>\n')
    f.write('<Class name="ArsJunctionArea" rules="update" traces="error">\n')
    f.write('      <Objects>\n')
    f.write('        <Object name="Jn_'+str(u)+'_PT01'+str(sh1.cell(row=1,column=2).value)+'" rules="update" traces="error">\n')
    f.write('          <Properties>\n')
    f.write('            <Property dt="string" name="AreaGroup">Area/LI_1/MI_1/Territory_'+str(u)+'/Area_'+str(u)+'</Property>\n')
    f.write('            <Property dt="string" name="FunctionGroup">Function/Signalling/ASCV/'+str(sh1.cell(row=1,column=2).value)+'</Property>\n')
    f.write('          </Properties>\n')
    f.write('        </Object>\n')
    f.write('      </Objects>\n')
    f.write('    </Class>\n')
    f.write('  </Classes>\n')



